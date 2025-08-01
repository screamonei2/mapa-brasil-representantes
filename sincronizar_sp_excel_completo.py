#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para sincronizar COMPLETAMENTE os dados de SP do Excel com o Sistema.
Adiciona SCHIOPPA 09.02 e corrige todas as diferenças encontradas.

MISSÃO 13 - FASE 1: Sincronização Excel → Sistema

Autor: Assistente AI
Data: $(date +%Y-%m-%d)
"""

import json
import openpyxl
import shutil
from datetime import datetime
from collections import defaultdict
import unicodedata
import re

def normalizar_nome_cidade(nome):
    """
    Normaliza nomes de cidades removendo acentos e caracteres especiais
    """
    nome = unicodedata.normalize('NFD', nome)
    nome = ''.join(char for char in nome if unicodedata.category(char) != 'Mn')
    nome = nome.upper()
    nome = re.sub(r'[^A-Z0-9\s\-\']', '', nome)
    nome = ' '.join(nome.split())
    return nome

def fazer_backup(arquivo):
    """
    Faz backup do arquivo original
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"{arquivo.replace('.json', '')}_backup_sync_sp_{timestamp}.json"
    shutil.copy2(arquivo, backup_name)
    print(f"✅ Backup criado: {backup_name}")
    return backup_name

def extrair_dados_excel():
    """
    Extrai dados completos do Excel para SP
    """
    print("📊 Extraindo dados do Excel...")
    
    wb = openpyxl.load_workbook('old/rep.xlsx')
    ws = wb.active
    
    representantes_excel = defaultdict(lambda: {
        'cidades': set(),
        'contato': None,
        'email': None,
        'celular': None,
        'nome': None
    })
    
    for row in range(2, ws.max_row + 1):
        sigla_estado = ws.cell(row=row, column=6).value
        
        if sigla_estado and str(sigla_estado).strip().upper() == 'SP':
            codigo = str(ws.cell(row=row, column=1).value).strip()
            nome_rep = str(ws.cell(row=row, column=2).value).strip()
            contato = ws.cell(row=row, column=3).value
            email = ws.cell(row=row, column=4).value
            celular = ws.cell(row=row, column=5).value
            cidade = str(ws.cell(row=row, column=8).value).strip().upper()
            
            if codigo and nome_rep and cidade:
                representantes_excel[codigo]['nome'] = nome_rep
                representantes_excel[codigo]['cidades'].add(cidade)
                if contato and not representantes_excel[codigo]['contato']:
                    representantes_excel[codigo]['contato'] = str(contato).strip()
                if email and not representantes_excel[codigo]['email']:
                    representantes_excel[codigo]['email'] = str(email).strip()
                if celular and not representantes_excel[codigo]['celular']:
                    representantes_excel[codigo]['celular'] = str(celular).strip()
    
    # Converter sets para listas ordenadas
    for codigo in representantes_excel:
        representantes_excel[codigo]['cidades'] = sorted(list(representantes_excel[codigo]['cidades']))
    
    print(f"✅ Extraídos dados de {len(representantes_excel)} representantes do Excel")
    return dict(representantes_excel)

def sincronizar_representantes():
    """
    Função principal para sincronizar representantes do Excel com o Sistema
    """
    arquivo_sistema = 'old/representantes_por_estado.json'
    
    print("🔄 SINCRONIZAÇÃO COMPLETA SP: Excel → Sistema")
    print("=" * 60)
    print()
    
    # 1. Fazer backup
    backup_file = fazer_backup(arquivo_sistema)
    
    # 2. Extrair dados do Excel
    representantes_excel = extrair_dados_excel()
    
    # 3. Carregar sistema atual
    print("📂 Carregando sistema atual...")
    with open(arquivo_sistema, 'r', encoding='utf-8') as f:
        data_sistema = json.load(f)
    
    print(f"✅ Sistema carregado com {len(data_sistema['representantes'])} representantes")
    print()
    
    # 4. Identificar diferenças
    print("🔍 ANÁLISE DE DIFERENÇAS:")
    print("-" * 40)
    
    representantes_adicionados = 0
    representantes_atualizados = 0
    cidades_adicionadas = 0
    
    for codigo_excel, dados_excel in representantes_excel.items():
        # Procurar representante no sistema pelo código
        rep_encontrado = None
        rep_key = None
        
        for key, rep in data_sistema['representantes'].items():
            if rep.get('codigo') == codigo_excel:
                rep_encontrado = rep
                rep_key = key
                break
        
        if rep_encontrado:
            # Representante existe - atualizar dados
            print(f"🔄 Atualizando {dados_excel['nome']} ({codigo_excel})")
            
            # Atualizar cidades de SP
            if 'SP' in rep_encontrado.get('estados', {}):
                cidades_atuais = set(rep_encontrado['estados']['SP']['cidades'])
                cidades_excel_set = set(dados_excel['cidades'])
                
                if cidades_atuais != cidades_excel_set:
                    novas_cidades = cidades_excel_set - cidades_atuais
                    cidades_removidas = cidades_atuais - cidades_excel_set
                    
                    print(f"   Cidades atuais: {len(cidades_atuais)}")
                    print(f"   Cidades Excel: {len(cidades_excel_set)}")
                    if novas_cidades:
                        print(f"   Adicionando: {sorted(list(novas_cidades))}")
                        cidades_adicionadas += len(novas_cidades)
                    if cidades_removidas:
                        print(f"   Removendo: {sorted(list(cidades_removidas))}")
                    
                    # Atualizar no sistema
                    rep_encontrado['estados']['SP']['cidades'] = dados_excel['cidades']
                    rep_encontrado['estados']['SP']['total_cidades'] = len(dados_excel['cidades'])
                    
                    # Atualizar total geral se só atende SP
                    if len(rep_encontrado['estados']) == 1:
                        rep_encontrado['total_cidades'] = len(dados_excel['cidades'])
                    
                    representantes_atualizados += 1
                else:
                    print(f"   ✅ Já sincronizado")
            else:
                # Adicionar SP ao representante existente
                print(f"   ➕ Adicionando SP ao representante")
                rep_encontrado['estados']['SP'] = {
                    'cidades': dados_excel['cidades'],
                    'total_cidades': len(dados_excel['cidades'])
                }
                if 'SP   ' not in rep_encontrado['estados_atendidos']:
                    rep_encontrado['estados_atendidos'].append('SP   ')
                cidades_adicionadas += len(dados_excel['cidades'])
                representantes_atualizados += 1
        else:
            # Representante NÃO existe - criar novo
            print(f"➕ CRIANDO NOVO: {dados_excel['nome']} ({codigo_excel})")
            print(f"   Cidades: {len(dados_excel['cidades'])}")
            print(f"   Região: {dados_excel['cidades'][:5]}...")
            
            # Criar chave única
            nome_key = dados_excel['nome'].lower().replace(' ', ' ').replace('ltda', 'ltda').replace('  ', ' ')
            nome_key = re.sub(r'[^a-z0-9\s]', '', nome_key)
            nome_key = ' '.join(nome_key.split())
            
            # Garantir que a chave seja única
            base_key = nome_key
            counter = 1
            while nome_key in data_sistema['representantes']:
                nome_key = f"{base_key} {counter}"
                counter += 1
            
            # Criar representante
            novo_representante = {
                "codigo": codigo_excel,
                "nome": dados_excel['nome'],
                "contato": {
                    "nome_contato": dados_excel['contato'] or "Não informado",
                    "email": dados_excel['email'] or "Não informado",
                    "celular": dados_excel['celular'] or "Não informado"
                },
                "observacoes": "Será enviado para pesquisa no mapa",
                "total_cidades": len(dados_excel['cidades']),
                "estados_atendidos": ["SP   "],
                "resumo_atividades": "",
                "performance": {},
                "estados": {
                    "SP": {
                        "cidades": dados_excel['cidades'],
                        "total_cidades": len(dados_excel['cidades'])
                    }
                }
            }
            
            data_sistema['representantes'][nome_key] = novo_representante
            representantes_adicionados += 1
            cidades_adicionadas += len(dados_excel['cidades'])
        
        print()
    
    # 5. Salvar arquivo atualizado
    print("💾 Salvando sistema atualizado...")
    with open(arquivo_sistema, 'w', encoding='utf-8') as f:
        json.dump(data_sistema, f, indent=2, ensure_ascii=False)
    
    # 6. Relatório final
    print("=" * 60)
    print("📊 RELATÓRIO DE SINCRONIZAÇÃO:")
    print(f"✅ Representantes adicionados: {representantes_adicionados}")
    print(f"🔄 Representantes atualizados: {representantes_atualizados}")
    print(f"➕ Cidades adicionadas ao sistema: {cidades_adicionadas}")
    print(f"💾 Backup salvo: {backup_file}")
    print(f"📁 Sistema atualizado: {arquivo_sistema}")
    print()
    print("🎯 FASE 1 CONCLUÍDA!")
    print("Próximo: FASE 2 - Correções de dados")
    
    return True

if __name__ == "__main__":
    try:
        sucesso = sincronizar_representantes()
        if sucesso:
            print("✅ Sincronização concluída com sucesso!")
        else:
            print("❌ Problemas na sincronização.")
    except Exception as e:
        print(f"❌ ERRO durante a sincronização: {e}")
        print("Verifique os arquivos e tente novamente.")